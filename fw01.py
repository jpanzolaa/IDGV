import shutil
import openpyxl
import IDGV

# 1. Crear copia del archivo
archivo_trabajo = "mnt/data/wp01E.xlsx"
shutil.copy("mnt/data/data0.xlsx", archivo_trabajo)

# 2. Abrir el archivo de trabajo y establecer encabezados
wb = openpyxl.load_workbook(archivo_trabajo)
hoja = wb.active

# Escribir nombres de columna desde la columna C en adelante
columnas = ["Ins Text", "Tox_a", "Sent", "Conf_Sent", "Pers", "Tox_b", "Ins", "eps", "IDVG"]
for i, nombre in enumerate(columnas, start=3):  # desde la columna C (índice 3)
    hoja.cell(row=1, column=i, value=nombre)

# 3. Inicializar variable registro
registro = 2

# 6. Determinar cantidad total de registros a procesar
total_filas = hoja.max_row

# 4. Bucle de procesamiento por fila
while registro <= total_filas:
    texto = hoja[f"A{registro}"].value
    if texto is None:
        break

    try:
        res = IDGV.calcular_idvg(texto)

        # 4. Escribir los resultados en las columnas correspondientes
        hoja[f"C{registro}"].value = str(res["Ins Text"])
        hoja[f"D{registro}"].value = res["Tox_a"]
        hoja[f"E{registro}"].value = res["Sent"]
        hoja[f"F{registro}"].value = res["Conf_Sent"]
        hoja[f"G{registro}"].value = res["Pers"]
        hoja[f"H{registro}"].value = res["Tox_b"]
        hoja[f"I{registro}"].value = res["Ins"]
        hoja[f"J{registro}"].value = res["eps"]
        hoja[f"K{registro}"].value = res["IDVG"]

        # 5. Guardar cambios
        wb.save(archivo_trabajo)

        # 6. Imprimir mensaje de control
        print(f"Registro {registro} de {total_filas} procesado.")

    except Exception as e:
        print(f"Error en el registro {registro}: {e}")

    # 7. Incrementar registro
    registro += 1

# 8. Mensaje final
print("Proceso finalizado.")
