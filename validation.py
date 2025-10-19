import os
import re
import zipfile
from pathlib import Path
from openpyxl import Workbook
import openpyxl
import shutil
import IDGV

# =========================
# 0. Configuración de rutas
# =========================
zip_path = r"mnt\data\Chat_validation.zip"              # ZIP de entrada
extract_folder = r"mnt\data\extracted"                  # Carpeta de extracción
out_dir = Path("mnt/data/test_validation")              # Carpeta del archivo de trabajo
out_dir.mkdir(parents=True, exist_ok=True)
archivo_trabajo = str(out_dir / "test.xlsx")            # Archivo de trabajo de salida

# ======================================================
# 1. Descomprimir ZIP y leer el primer archivo de texto
# ======================================================
os.makedirs(extract_folder, exist_ok=True)

with zipfile.ZipFile(zip_path, 'r') as zip_ref:
    zip_ref.extractall(extract_folder)
    print("Archivos extraídos en:", extract_folder)
    extracted_files = zip_ref.namelist()
    print("Archivos encontrados en el ZIP:", extracted_files)

# Selecciona el primer .txt (ajusta si necesitas otra lógica)
txt_path = None
for name in extracted_files:
    if name.lower().endswith(".txt"):
        txt_path = os.path.join(extract_folder, name)
        break

if not txt_path or not os.path.exists(txt_path):
    raise FileNotFoundError("No se encontró un archivo .txt dentro del ZIP.")

# Lee el contenido (maneja BOM si existiera)
with open(txt_path, "r", encoding="utf-8-sig") as f:
    content = f.read()

print("\nContenido del archivo de texto:")
print(content)

# =====================================================================
# 2. Parseo de 'content' para extraer mensajes posteriores al aviso E2E
# =====================================================================
# Aviso de cifrado
aviso_regex = re.compile(
    r"Los mensajes y las llamadas están cifrados de extremo a extremo\..*?Obtén más información\.",
    flags=re.DOTALL | re.UNICODE
)

split_match = aviso_regex.search(content)
if split_match:
    mensajes_raw = content[split_match.end():].strip()
else:
    mensajes_raw = content.strip()

# Normalización de espacios: reemplaza NBSP y NNBSP por espacio normal y compacta
def normalize(s: str) -> str:
    s = s.replace("\u00A0", " ").replace("\u202F", " ")
    s = re.sub(r"[ \t]+", " ", s, flags=re.UNICODE)
    return s

mensajes_raw = normalize(mensajes_raw)

# Regex para nuevas líneas de mensaje tipo WhatsApp
# Fecha: dd/mm/yy
# Hora:  h:mm a. m. / p. m. (con espacios/puntos variables)
# Sujeto: entre " - " y ":"
pat_inicio = re.compile(
    r"(?P<fecha>\d{1,2}/\d{1,2}/\d{2})\s*,\s*"
    r"(?P<hora>\d{1,2}:\d{2}\s*[ap]\.?\s*m\.?)\s*-\s*"
    r"(?P<sujeto>[^:]+)\s*:\s*"
    r"(?P<mensaje>.*)",
    flags=re.IGNORECASE | re.UNICODE
)

rows = []  # (fecha, hora, sujeto, mensaje)
current = None

for line in mensajes_raw.splitlines():
    raw = line.strip()                 # <= elimina \r y espacios a ambos lados
    if not raw:
        continue
    m = pat_inicio.match(raw)
    if m:
        # guarda el anterior
        if current:
            current["mensaje"] = current["mensaje"].strip()
            rows.append((current["fecha"], current["hora"], current["sujeto"], current["mensaje"]))
        # inicia nuevo
        current = {
            "fecha": m.group("fecha").strip(),
            "hora": normalize(m.group("hora")),
            "sujeto": m.group("sujeto").strip(),
            "mensaje": m.group("mensaje")
        }
    else:
        # línea de continuación
        if current:
            current["mensaje"] += (" " if current["mensaje"] else "") + raw

# cierra el último
if current:
    current["mensaje"] = current["mensaje"].strip()
    rows.append((current["fecha"], current["hora"], current["sujeto"], current["mensaje"]))

# ===========================================================
# 3. Crear archivo de trabajo Excel y escribir encabezados y filas
# ===========================================================
if os.path.exists(archivo_trabajo):
    wb = openpyxl.load_workbook(archivo_trabajo)
    hoja = wb.active
else:
    wb = Workbook()
    hoja = wb.active

# Encabezados exactos
columnas = ["Fecha", "Hora", "Sujeto", "mensaje ", "Ins Text", "Tox_a", "Sent", "Conf_Sent",
            "Pers", "Tox_b", "Ins", "eps", "IDVG"]

# Limpia hoja
hoja.delete_rows(1, hoja.max_row)

# Escribe encabezados
for idx, col_name in enumerate(columnas, start=1):
    hoja.cell(row=1, column=idx, value=col_name)

# Escribe filas
start_row = 2
for i, (fecha, hora, sujeto, mensaje) in enumerate(rows, start=start_row):
    hoja.cell(row=i, column=1, value=fecha)    # A
    hoja.cell(row=i, column=2, value=hora)     # B
    hoja.cell(row=i, column=3, value=sujeto)   # C
    hoja.cell(row=i, column=4, value=mensaje)  # D

wb.save(archivo_trabajo)
print(f"\nArchivo de trabajo creado/actualizado: {archivo_trabajo}")
print(f"Total de mensajes escritos: {len(rows)}")

# Depuración extra útil
if not rows:
    print("ADVERTENCIA: No se detectaron mensajes. Revisa el formato de fecha/hora y el aviso de cifrado.")

# 2. Abrir el archivo de trabajo y establecer encabezados
wb = openpyxl.load_workbook(archivo_trabajo)
hoja = wb.active

# 3. Inicializar variable registro
registro = 2

# 6. Determinar cantidad total de registros a procesar
total_filas = hoja.max_row

# 4. Bucle de procesamiento por fila
while registro <= total_filas:
    texto = hoja[f"D{registro}"].value
    if texto is None:
        break

    try:
        res = IDGV.calcular_idvg(texto)

        # 4. Escribir los resultados en las columnas correspondientes
        hoja[f"E{registro}"].value = str(res["Ins Text"])
        hoja[f"F{registro}"].value = res["Tox_a"]
        hoja[f"G{registro}"].value = res["Sent"]
        hoja[f"H{registro}"].value = res["Conf_Sent"]
        hoja[f"I{registro}"].value = res["Pers"]
        hoja[f"J{registro}"].value = res["Tox_b"]
        hoja[f"K{registro}"].value = res["Ins"]
        hoja[f"L{registro}"].value = res["eps"]
        hoja[f"M{registro}"].value = res["IDVG"]

        # 5. Guardar cambios
        wb.save(archivo_trabajo)

        # 6. Imprimir mensaje de control
        print(f"Registro {registro} de {total_filas} procesado.")

    except Exception as e:
        print(f"Error en el registro {registro}: {e}")

    # 7. Incrementar registro
    registro += 1

# 8. Mensaje final
print("Proceso finalizado.")

